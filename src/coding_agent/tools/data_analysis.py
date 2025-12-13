"""Data analysis tools.

These tools provide structured, higher-level data analysis capabilities beyond the general `python_repl` sandbox.
They intentionally avoid arbitrary code execution and instead expose common operations as explicit tools.

Supported formats:
- CSV (`.csv`)
- JSON (`.json`) - list of objects, or dict containing a list under `data`/`records`
- JSON Lines (`.jsonl`)
- Excel (`.xlsx`) via optional dependency `openpyxl`

Most tools operate on an in-memory dataset registry keyed by `dataset_id`.
"""

from __future__ import annotations

import csv
import json
import math
import random
import statistics
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from ..exceptions import PathTraversalError
from .base import BaseTool
from .filesystem import get_path_validator

_MISSING_SENTINELS = {"", "na", "n/a", "null", "none", "nan"}


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip().lower() in _MISSING_SENTINELS:
        return True
    return False


def _coerce_float(value: Any) -> float | None:
    if _is_missing(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return float(value)
        except Exception:
            return None
    if isinstance(value, str):
        txt = value.strip().replace(",", "")
        if not txt:
            return None
        try:
            return float(txt)
        except Exception:
            return None
    return None


def _infer_type(values: Iterable[Any]) -> str:
    seen_numeric = False
    seen_text = False
    for v in values:
        if _is_missing(v):
            continue
        if _coerce_float(v) is not None:
            seen_numeric = True
        else:
            seen_text = True
        if seen_numeric and seen_text:
            return "mixed"
    if seen_numeric:
        return "numeric"
    if seen_text:
        return "text"
    return "empty"


def _format_value(value: Any) -> str:
    if _is_missing(value):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _render_markdown_table(rows: list[dict[str, Any]], columns: list[str], max_rows: int = 20) -> str:
    cols = columns
    shown = rows[: max_rows]
    header = "| " + " | ".join(cols) + " |"
    sep = "| " + " | ".join(["---"] * len(cols)) + " |"
    lines = [header, sep]
    for row in shown:
        lines.append("| " + " | ".join(_format_value(row.get(c)) for c in cols) + " |")
    if len(rows) > max_rows:
        lines.append(f"\n… ({len(rows) - max_rows} more rows)")
    return "\n".join(lines)


def _new_dataset_id() -> str:
    return uuid.uuid4().hex[:10]


@dataclass(frozen=True)
class Dataset:
    id: str
    name: str
    columns: list[str]
    rows: list[dict[str, Any]]
    source: str | None = None


_DATASETS: dict[str, Dataset] = {}


def _get_dataset(dataset_id: str) -> Dataset | None:
    return _DATASETS.get(dataset_id)


def _resolve_path(path: str) -> Path:
    validator = get_path_validator()
    return validator.validate(path)


def _ensure_parent_dir(path: Path) -> None:
    parent = path.parent
    if not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)


def _auto_format_from_path(path: str) -> str:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix == ".json":
        return "json"
    if suffix == ".jsonl":
        return "jsonl"
    if suffix == ".xlsx":
        return "xlsx"
    return ""


def _load_csv(resolved: Path, *, delimiter: str, encoding: str, max_rows: int | None) -> tuple[list[str], list[dict[str, Any]]]:
    with open(resolved, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        fieldnames = reader.fieldnames or []
        rows: list[dict[str, Any]] = []
        for i, row in enumerate(reader):
            if max_rows is not None and i >= max_rows:
                break
            rows.append(dict(row))
    return fieldnames, rows


def _load_json(resolved: Path, *, encoding: str, max_rows: int | None) -> tuple[list[str], list[dict[str, Any]]]:
    with open(resolved, "r", encoding=encoding) as f:
        data = json.load(f)

    records: list[dict[str, Any]]
    if isinstance(data, list):
        records = [r for r in data if isinstance(r, dict)]
    elif isinstance(data, dict):
        candidate = None
        for key in ("data", "records", "items"):
            if isinstance(data.get(key), list):
                candidate = data[key]
                break
        if isinstance(candidate, list):
            records = [r for r in candidate if isinstance(r, dict)]
        else:
            records = [data]
    else:
        records = [{"value": data}]

    if max_rows is not None:
        records = records[:max_rows]

    columns: list[str] = []
    seen = set()
    for r in records:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                columns.append(str(k))

    normalized: list[dict[str, Any]] = []
    for r in records:
        normalized.append({str(k): v for k, v in r.items()})

    return columns, normalized


def _load_jsonl(resolved: Path, *, encoding: str, max_rows: int | None) -> tuple[list[str], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    with open(resolved, "r", encoding=encoding) as f:
        for i, line in enumerate(f):
            if max_rows is not None and i >= max_rows:
                break
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(obj, dict):
                rows.append({str(k): v for k, v in obj.items()})
            else:
                rows.append({"value": obj})

    columns: list[str] = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)
    return columns, rows


def _load_xlsx(
    resolved: Path,
    *,
    sheet_name: str | None,
    sheet_index: int | None,
    has_header: bool,
    max_rows: int | None,
) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:
        raise RuntimeError("openpyxl is required for .xlsx support. Install with: uv sync --extra data")

    wb = load_workbook(resolved, read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            ws = wb[sheet_name]
        elif sheet_index is not None:
            ws = wb.worksheets[sheet_index]
        else:
            ws = wb.worksheets[0]

        iterator = ws.iter_rows(values_only=True)
        rows: list[tuple[Any, ...]] = []
        for i, r in enumerate(iterator):
            if max_rows is not None and i >= max_rows:
                break
            rows.append(tuple(r))

        if not rows:
            return [], []

        if has_header:
            raw_cols = rows[0]
            columns = [str(c) if c is not None else f"col_{idx+1}" for idx, c in enumerate(raw_cols)]
            data_rows = rows[1:]
        else:
            width = max(len(r) for r in rows)
            columns = [f"col_{idx+1}" for idx in range(width)]
            data_rows = rows

        normalized: list[dict[str, Any]] = []
        for r in data_rows:
            row_dict: dict[str, Any] = {}
            for idx, col in enumerate(columns):
                row_dict[col] = r[idx] if idx < len(r) else None
            normalized.append(row_dict)

        return columns, normalized
    finally:
        try:
            wb.close()
        except Exception:
            pass


class LoadDatasetTool(BaseTool):
    """Load a dataset from disk into the in-memory registry."""

    @property
    def name(self) -> str:
        return "load_dataset"

    @property
    def description(self) -> str:
        return (
            "Load a dataset from a file path into memory and return a dataset_id. "
            "Supports csv, json, jsonl, and xlsx."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to the dataset file (csv/json/jsonl/xlsx)."},
                "format": {
                    "type": "string",
                    "enum": ["auto", "csv", "json", "jsonl", "xlsx"],
                    "description": "File format. Use 'auto' to infer from extension.",
                    "default": "auto",
                },
                "name": {"type": "string", "description": "Optional human-friendly dataset name."},
                "delimiter": {"type": "string", "description": "CSV delimiter (only for csv).", "default": ","},
                "encoding": {"type": "string", "description": "Text encoding for csv/json/jsonl.", "default": "utf-8"},
                "max_rows": {"type": "integer", "description": "Optional max rows to load (for sampling)."},
                "sheet_name": {"type": "string", "description": "XLSX: sheet name to load."},
                "sheet_index": {"type": "integer", "description": "XLSX: 0-based sheet index to load."},
                "has_header": {
                    "type": "boolean",
                    "description": "XLSX: whether first row is header.",
                    "default": True,
                },
            },
            "required": ["path"],
        }

    def execute(  # noqa: PLR0913
        self,
        path: str,
        format: str = "auto",
        name: str | None = None,
        delimiter: str = ",",
        encoding: str = "utf-8",
        max_rows: int | None = None,
        sheet_name: str | None = None,
        sheet_index: int | None = None,
        has_header: bool = True,
    ) -> str:
        try:
            resolved = _resolve_path(path)
        except PathTraversalError as e:
            return f"Security error: {e}"

        fmt = _auto_format_from_path(path) if format == "auto" else format
        if fmt not in {"csv", "json", "jsonl", "xlsx"}:
            return f"Error: Unsupported format '{format}'. Supported: csv, json, jsonl, xlsx."

        try:
            if fmt == "csv":
                columns, rows = _load_csv(resolved, delimiter=delimiter, encoding=encoding, max_rows=max_rows)
            elif fmt == "json":
                columns, rows = _load_json(resolved, encoding=encoding, max_rows=max_rows)
            elif fmt == "jsonl":
                columns, rows = _load_jsonl(resolved, encoding=encoding, max_rows=max_rows)
            else:
                columns, rows = _load_xlsx(
                    resolved,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    has_header=has_header,
                    max_rows=max_rows,
                )
        except Exception as e:
            return f"Error: Failed to load dataset: {type(e).__name__}: {e}"

        dataset_id = _new_dataset_id()
        dataset_name = name or Path(path).name
        _DATASETS[dataset_id] = Dataset(
            id=dataset_id,
            name=dataset_name,
            columns=columns,
            rows=rows,
            source=str(resolved),
        )

        return (
            f"Loaded dataset '{dataset_name}' as dataset_id={dataset_id} "
            f"({len(rows)} rows, {len(columns)} columns)."
        )


class ListDatasetsTool(BaseTool):
    @property
    def name(self) -> str:
        return "list_datasets"

    @property
    def description(self) -> str:
        return "List datasets currently loaded in memory."

    @property
    def parameters(self) -> dict[str, Any]:
        return {"type": "object", "properties": {}, "required": []}

    def execute(self) -> str:
        if not _DATASETS:
            return "No datasets loaded."
        lines = []
        for ds in _DATASETS.values():
            lines.append(
                f"- {ds.id}: {ds.name} ({len(ds.rows)} rows, {len(ds.columns)} cols)"
            )
        return "\n".join(lines)


class RemoveDatasetTool(BaseTool):
    @property
    def name(self) -> str:
        return "remove_dataset"

    @property
    def description(self) -> str:
        return "Remove a dataset from memory by dataset_id."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {"dataset_id": {"type": "string", "description": "Dataset ID to remove."}},
            "required": ["dataset_id"],
        }

    def execute(self, dataset_id: str) -> str:
        if dataset_id not in _DATASETS:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        removed = _DATASETS.pop(dataset_id)
        return f"Removed dataset '{removed.name}' (dataset_id={dataset_id})."


class ClearDatasetsTool(BaseTool):
    @property
    def name(self) -> str:
        return "clear_datasets"

    @property
    def description(self) -> str:
        return "Clear all loaded datasets from memory."

    @property
    def parameters(self) -> dict[str, Any]:
        return {"type": "object", "properties": {}, "required": []}

    def execute(self) -> str:
        count = len(_DATASETS)
        _DATASETS.clear()
        return f"Cleared {count} dataset(s)."


class DatasetHeadTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_head"

    @property
    def description(self) -> str:
        return "Show the first N rows of a dataset as a markdown table."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "n": {"type": "integer", "description": "Number of rows to show.", "default": 10},
            },
            "required": ["dataset_id"],
        }

    def execute(self, dataset_id: str, n: int = 10) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        n = max(0, min(int(n), 50))
        return _render_markdown_table(ds.rows[:n], ds.columns, max_rows=n)


class DatasetTailTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_tail"

    @property
    def description(self) -> str:
        return "Show the last N rows of a dataset as a markdown table."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "n": {"type": "integer", "description": "Number of rows to show.", "default": 10},
            },
            "required": ["dataset_id"],
        }

    def execute(self, dataset_id: str, n: int = 10) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        n = max(0, min(int(n), 50))
        return _render_markdown_table(ds.rows[-n:], ds.columns, max_rows=n)


class DatasetSampleTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_sample"

    @property
    def description(self) -> str:
        return "Randomly sample N rows from a dataset (without replacement)."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "n": {"type": "integer", "description": "Number of rows to sample.", "default": 10},
                "seed": {"type": "integer", "description": "Optional random seed."},
            },
            "required": ["dataset_id"],
        }

    def execute(self, dataset_id: str, n: int = 10, seed: int | None = None) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        n = max(0, min(int(n), 50))
        if seed is not None:
            random.seed(seed)
        if n >= len(ds.rows):
            sampled = ds.rows
        else:
            sampled = random.sample(ds.rows, k=n)
        return _render_markdown_table(sampled, ds.columns, max_rows=n)


class DatasetInfoTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_info"

    @property
    def description(self) -> str:
        return "Show dataset size, column types, and missing value counts."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {"dataset_id": {"type": "string", "description": "Dataset ID."}},
            "required": ["dataset_id"],
        }

    def execute(self, dataset_id: str) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."

        sample_rows = ds.rows[: min(200, len(ds.rows))]
        col_values: dict[str, list[Any]] = {c: [] for c in ds.columns}
        missing: dict[str, int] = {c: 0 for c in ds.columns}

        for r in ds.rows:
            for c in ds.columns:
                v = r.get(c)
                if _is_missing(v):
                    missing[c] += 1
                if len(col_values[c]) < 200:
                    col_values[c].append(v)

        lines = [
            f"dataset_id: {ds.id}",
            f"name: {ds.name}",
            f"source: {ds.source or '(unknown)'}",
            f"rows: {len(ds.rows)}",
            f"columns: {len(ds.columns)}",
            "",
            "columns:",
        ]
        for c in ds.columns:
            inferred = _infer_type(col_values[c])
            lines.append(f"- {c}: type={inferred}, missing={missing[c]}")
        if len(sample_rows) != len(ds.rows):
            lines.append(f"\n(type inference based on first {len(sample_rows)} rows)")
        return "\n".join(lines)


class DatasetDescribeTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_describe"

    @property
    def description(self) -> str:
        return "Compute basic descriptive statistics for numeric columns (count, mean, stdev, min, p25, p50, p75, max)."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional list of numeric columns to describe (default: all).",
                },
            },
            "required": ["dataset_id"],
        }

    def execute(self, dataset_id: str, columns: list[str] | None = None) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."

        target_cols = columns or ds.columns
        stats_rows: list[dict[str, Any]] = []

        for c in target_cols:
            values = [_coerce_float(r.get(c)) for r in ds.rows]
            numeric = [v for v in values if v is not None]
            if not numeric:
                continue
            numeric.sort()
            n = len(numeric)
            mean = statistics.fmean(numeric)
            stdev = statistics.pstdev(numeric) if n > 1 else 0.0

            def percentile(p: float) -> float:
                if n == 1:
                    return numeric[0]
                idx = (n - 1) * p
                lo = int(math.floor(idx))
                hi = int(math.ceil(idx))
                if lo == hi:
                    return numeric[lo]
                w = idx - lo
                return numeric[lo] * (1 - w) + numeric[hi] * w

            stats_rows.append({
                "column": c,
                "count": n,
                "mean": mean,
                "stdev": stdev,
                "min": numeric[0],
                "p25": percentile(0.25),
                "p50": percentile(0.50),
                "p75": percentile(0.75),
                "max": numeric[-1],
            })

        if not stats_rows:
            return "No numeric columns found (or selected columns have no numeric values)."

        cols = ["column", "count", "mean", "stdev", "min", "p25", "p50", "p75", "max"]
        return _render_markdown_table(stats_rows, cols, max_rows=50)


class DatasetValueCountsTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_value_counts"

    @property
    def description(self) -> str:
        return "Compute value counts for a column (top N)."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "column": {"type": "string", "description": "Column name."},
                "top_n": {"type": "integer", "description": "Number of top values to show.", "default": 20},
                "include_missing": {"type": "boolean", "description": "Include missing as a category.", "default": True},
            },
            "required": ["dataset_id", "column"],
        }

    def execute(self, dataset_id: str, column: str, top_n: int = 20, include_missing: bool = True) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        if column not in ds.columns:
            return f"Error: Unknown column '{column}'. Available: {', '.join(ds.columns)}"

        counts: dict[str, int] = {}
        for r in ds.rows:
            v = r.get(column)
            if _is_missing(v):
                if not include_missing:
                    continue
                key = "(missing)"
            else:
                key = _format_value(v)
            counts[key] = counts.get(key, 0) + 1

        items = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)
        top_n = max(1, min(int(top_n), 50))
        rows = [{"value": k, "count": v} for k, v in items[:top_n]]
        return _render_markdown_table(rows, ["value", "count"], max_rows=top_n)


class DatasetSelectColumnsTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_select_columns"

    @property
    def description(self) -> str:
        return "Create a new dataset with only selected columns."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "columns": {"type": "array", "items": {"type": "string"}, "description": "Columns to keep."},
                "name": {"type": "string", "description": "Optional new dataset name."},
            },
            "required": ["dataset_id", "columns"],
        }

    def execute(self, dataset_id: str, columns: list[str], name: str | None = None) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        missing = [c for c in columns if c not in ds.columns]
        if missing:
            return f"Error: Unknown columns: {', '.join(missing)}"

        new_rows = [{c: r.get(c) for c in columns} for r in ds.rows]
        new_id = _new_dataset_id()
        _DATASETS[new_id] = Dataset(
            id=new_id,
            name=name or f"{ds.name} (select)",
            columns=list(columns),
            rows=new_rows,
            source=ds.source,
        )
        return f"Created dataset_id={new_id} with {len(new_rows)} rows and {len(columns)} columns."


class DatasetFilterTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_filter"

    @property
    def description(self) -> str:
        return "Filter rows by simple conditions and return a new dataset_id."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "conditions": {
                    "type": "array",
                    "description": "List of conditions; all must match (AND).",
                    "items": {
                        "type": "object",
                        "properties": {
                            "column": {"type": "string"},
                            "op": {
                                "type": "string",
                                "enum": ["==", "!=", ">", ">=", "<", "<=", "contains", "startswith", "endswith"],
                            },
                            "value": {"type": ["string", "number", "boolean"]},
                        },
                        "required": ["column", "op", "value"],
                    },
                },
                "name": {"type": "string", "description": "Optional new dataset name."},
            },
            "required": ["dataset_id", "conditions"],
        }

    def execute(self, dataset_id: str, conditions: list[dict[str, Any]], name: str | None = None) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."

        for cond in conditions:
            col = cond.get("column")
            if col not in ds.columns:
                return f"Error: Unknown column '{col}'. Available: {', '.join(ds.columns)}"

        def matches(row: dict[str, Any]) -> bool:
            for cond in conditions:
                col = str(cond["column"])
                op = str(cond["op"])
                target = cond["value"]
                val = row.get(col)

                if op in {">", ">=", "<", "<="}:
                    a = _coerce_float(val)
                    b = _coerce_float(target)
                    if a is None or b is None:
                        return False
                    if op == ">" and not (a > b):
                        return False
                    if op == ">=" and not (a >= b):
                        return False
                    if op == "<" and not (a < b):
                        return False
                    if op == "<=" and not (a <= b):
                        return False
                    continue

                if op == "==":
                    if _format_value(val) != _format_value(target):
                        return False
                elif op == "!=":
                    if _format_value(val) == _format_value(target):
                        return False
                elif op == "contains":
                    if _format_value(target) not in _format_value(val):
                        return False
                elif op == "startswith":
                    if not _format_value(val).startswith(_format_value(target)):
                        return False
                elif op == "endswith":
                    if not _format_value(val).endswith(_format_value(target)):
                        return False
                else:
                    return False
            return True

        filtered = [r for r in ds.rows if matches(r)]
        new_id = _new_dataset_id()
        _DATASETS[new_id] = Dataset(
            id=new_id,
            name=name or f"{ds.name} (filtered)",
            columns=ds.columns,
            rows=filtered,
            source=ds.source,
        )
        return f"Created dataset_id={new_id} with {len(filtered)} rows (from {len(ds.rows)})."


class DatasetSortTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_sort"

    @property
    def description(self) -> str:
        return "Sort a dataset by one or more columns and return a new dataset_id."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "by": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Columns to sort by (in priority order).",
                },
                "ascending": {"type": "boolean", "description": "Sort ascending (default: true).", "default": True},
                "name": {"type": "string", "description": "Optional new dataset name."},
            },
            "required": ["dataset_id", "by"],
        }

    def execute(self, dataset_id: str, by: list[str], ascending: bool = True, name: str | None = None) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."

        missing = [c for c in by if c not in ds.columns]
        if missing:
            return f"Error: Unknown columns: {', '.join(missing)}"

        def sort_key(row: dict[str, Any]) -> tuple:
            key_parts: list[tuple[int, Any]] = []
            for col in by:
                v = row.get(col)
                num = _coerce_float(v)
                if num is not None:
                    key_parts.append((0, num))
                else:
                    key_parts.append((1, _format_value(v)))
            return tuple(key_parts)

        new_rows = sorted(ds.rows, key=sort_key, reverse=not ascending)
        new_id = _new_dataset_id()
        _DATASETS[new_id] = Dataset(
            id=new_id,
            name=name or f"{ds.name} (sorted)",
            columns=ds.columns,
            rows=new_rows,
            source=ds.source,
        )
        return f"Created dataset_id={new_id} with {len(new_rows)} rows (sorted)."


class DatasetGroupByAggTool(BaseTool):
    @property
    def name(self) -> str:
        return "dataset_groupby_agg"

    @property
    def description(self) -> str:
        return "Group by columns and compute an aggregation (count/sum/mean/min/max) over a numeric column."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "group_by": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Columns to group by.",
                },
                "agg": {
                    "type": "string",
                    "enum": ["count", "sum", "mean", "min", "max"],
                    "description": "Aggregation to compute.",
                    "default": "count",
                },
                "value_column": {
                    "type": "string",
                    "description": "Numeric column to aggregate (required unless agg=count).",
                },
                "name": {"type": "string", "description": "Optional new dataset name."},
            },
            "required": ["dataset_id", "group_by"],
        }

    def execute(
        self,
        dataset_id: str,
        group_by: list[str],
        agg: str = "count",
        value_column: str | None = None,
        name: str | None = None,
    ) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."

        missing = [c for c in group_by if c not in ds.columns]
        if missing:
            return f"Error: Unknown group_by columns: {', '.join(missing)}"

        if agg != "count":
            if not value_column:
                return "Error: value_column is required unless agg='count'."
            if value_column not in ds.columns:
                return f"Error: Unknown value_column '{value_column}'."

        groups: dict[tuple[str, ...], list[dict[str, Any]]] = {}
        for r in ds.rows:
            key = tuple(_format_value(r.get(c)) for c in group_by)
            groups.setdefault(key, []).append(r)

        out_rows: list[dict[str, Any]] = []
        out_col = "count" if agg == "count" else f"{agg}({value_column})"

        for key, items in groups.items():
            out: dict[str, Any] = {c: v for c, v in zip(group_by, key, strict=False)}
            if agg == "count":
                out[out_col] = len(items)
            else:
                vals = [_coerce_float(r.get(value_column or "")) for r in items]
                numeric = [v for v in vals if v is not None]
                if not numeric:
                    out[out_col] = None
                elif agg == "sum":
                    out[out_col] = sum(numeric)
                elif agg == "mean":
                    out[out_col] = statistics.fmean(numeric)
                elif agg == "min":
                    out[out_col] = min(numeric)
                elif agg == "max":
                    out[out_col] = max(numeric)
                else:
                    out[out_col] = None
            out_rows.append(out)

        new_id = _new_dataset_id()
        out_columns = list(group_by) + [out_col]
        _DATASETS[new_id] = Dataset(
            id=new_id,
            name=name or f"{ds.name} (groupby {agg})",
            columns=out_columns,
            rows=out_rows,
            source=ds.source,
        )
        return f"Created dataset_id={new_id} with {len(out_rows)} rows (grouped)."


class ExportDatasetTool(BaseTool):
    """Export a dataset to disk (csv/jsonl/xlsx)."""

    REQUIRES_CONFIRMATION = True
    OPERATION_TYPE = "write"
    CONFIRMATION_CHECK_ARG = "path"
    CONFIRMATION_MESSAGE = "Export dataset '{dataset_id}' to '{path}' as {format}"

    @property
    def name(self) -> str:
        return "export_dataset"

    @property
    def description(self) -> str:
        return (
            "Export a dataset to disk. Supports csv, jsonl, and xlsx. "
            "Ask the user for the output path before calling this tool."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "path": {"type": "string", "description": "Output file path."},
                "format": {"type": "string", "enum": ["csv", "jsonl", "xlsx"], "description": "Output format."},
                "overwrite": {"type": "boolean", "description": "Overwrite if file exists.", "default": False},
                "sheet_name": {"type": "string", "description": "XLSX: sheet name (default: Sheet1)."},
            },
            "required": ["dataset_id", "path", "format"],
        }

    def execute(
        self,
        dataset_id: str,
        path: str,
        format: str,
        overwrite: bool = False,
        sheet_name: str | None = None,
    ) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."

        try:
            out_path = _resolve_path(path)
        except PathTraversalError as e:
            return f"Security error: {e}"

        if out_path.exists() and not overwrite:
            return (
                f"Error: Output file already exists: {path}. "
                "Ask the user whether to overwrite, then retry with overwrite=true."
            )

        try:
            _ensure_parent_dir(out_path)
            if format == "csv":
                with open(out_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=ds.columns)
                    writer.writeheader()
                    for r in ds.rows:
                        writer.writerow({c: r.get(c) for c in ds.columns})
            elif format == "jsonl":
                with open(out_path, "w", encoding="utf-8") as f:
                    for r in ds.rows:
                        f.write(json.dumps(r, ensure_ascii=False))
                        f.write("\n")
            else:
                try:
                    from openpyxl import Workbook  # type: ignore[import-not-found]
                except Exception:
                    return "Error: openpyxl is required for xlsx export. Install with: uv sync --extra data"

                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name or "Sheet1"
                ws.append(list(ds.columns))
                for r in ds.rows:
                    ws.append([r.get(c) for c in ds.columns])
                wb.save(out_path)
                wb.close()
        except Exception as e:
            return f"Error: Failed to export dataset: {type(e).__name__}: {e}"

        return f"Exported dataset_id={dataset_id} to {path} ({format})."


class SaveHistogramPlotTool(BaseTool):
    REQUIRES_CONFIRMATION = True
    OPERATION_TYPE = "write"
    CONFIRMATION_CHECK_ARG = "path"
    CONFIRMATION_MESSAGE = "Save histogram plot to '{path}'"

    @property
    def name(self) -> str:
        return "save_histogram_plot"

    @property
    def description(self) -> str:
        return (
            "Save a histogram plot for a numeric column to an image file (png/svg). "
            "Ask the user for the output path before calling."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "column": {"type": "string", "description": "Numeric column."},
                "path": {"type": "string", "description": "Output image path (e.g., plots/hist.png)."},
                "bins": {"type": "integer", "description": "Histogram bins.", "default": 20},
                "title": {"type": "string", "description": "Optional plot title."},
                "overwrite": {"type": "boolean", "description": "Overwrite if file exists.", "default": False},
            },
            "required": ["dataset_id", "column", "path"],
        }

    def execute(
        self,
        dataset_id: str,
        column: str,
        path: str,
        bins: int = 20,
        title: str | None = None,
        overwrite: bool = False,
    ) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        if column not in ds.columns:
            return f"Error: Unknown column '{column}'. Available: {', '.join(ds.columns)}"

        try:
            out_path = _resolve_path(path)
        except PathTraversalError as e:
            return f"Security error: {e}"

        if out_path.exists() and not overwrite:
            return (
                f"Error: Output file already exists: {path}. "
                "Ask the user whether to overwrite, then retry with overwrite=true."
            )

        values = [_coerce_float(r.get(column)) for r in ds.rows]
        numeric = [v for v in values if v is not None]
        if not numeric:
            return f"Error: Column '{column}' has no numeric values."

        try:
            import matplotlib  # type: ignore[import-not-found]

            matplotlib.use("Agg")  # non-interactive backend
            import matplotlib.pyplot as plt  # type: ignore[import-not-found]
        except Exception:
            return "Error: matplotlib is required for plot saving. Install with: uv sync --extra data"

        try:
            _ensure_parent_dir(out_path)
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.hist(numeric, bins=max(1, int(bins)))
            ax.set_xlabel(column)
            ax.set_ylabel("count")
            ax.set_title(title or f"Histogram: {column}")
            fig.tight_layout()
            fig.savefig(out_path)
            plt.close(fig)
        except Exception as e:
            return f"Error: Failed to save plot: {type(e).__name__}: {e}"

        return f"Saved histogram plot to {path}."


class SaveScatterPlotTool(BaseTool):
    REQUIRES_CONFIRMATION = True
    OPERATION_TYPE = "write"
    CONFIRMATION_CHECK_ARG = "path"
    CONFIRMATION_MESSAGE = "Save scatter plot to '{path}'"

    @property
    def name(self) -> str:
        return "save_scatter_plot"

    @property
    def description(self) -> str:
        return (
            "Save a scatter plot for two numeric columns to an image file (png/svg). "
            "Ask the user for the output path before calling."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "x": {"type": "string", "description": "X column (numeric)."},
                "y": {"type": "string", "description": "Y column (numeric)."},
                "path": {"type": "string", "description": "Output image path."},
                "title": {"type": "string", "description": "Optional plot title."},
                "overwrite": {"type": "boolean", "description": "Overwrite if file exists.", "default": False},
            },
            "required": ["dataset_id", "x", "y", "path"],
        }

    def execute(
        self,
        dataset_id: str,
        x: str,
        y: str,
        path: str,
        title: str | None = None,
        overwrite: bool = False,
    ) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        for col in (x, y):
            if col not in ds.columns:
                return f"Error: Unknown column '{col}'. Available: {', '.join(ds.columns)}"

        try:
            out_path = _resolve_path(path)
        except PathTraversalError as e:
            return f"Security error: {e}"

        if out_path.exists() and not overwrite:
            return (
                f"Error: Output file already exists: {path}. "
                "Ask the user whether to overwrite, then retry with overwrite=true."
            )

        xs = [_coerce_float(r.get(x)) for r in ds.rows]
        ys = [_coerce_float(r.get(y)) for r in ds.rows]
        points = [(a, b) for a, b in zip(xs, ys, strict=False) if a is not None and b is not None]
        if not points:
            return f"Error: No numeric pairs found for columns '{x}' and '{y}'."

        try:
            import matplotlib  # type: ignore[import-not-found]

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt  # type: ignore[import-not-found]
        except Exception:
            return "Error: matplotlib is required for plot saving. Install with: uv sync --extra data"

        try:
            _ensure_parent_dir(out_path)
            fig, ax = plt.subplots(figsize=(7, 5))
            ax.scatter([p[0] for p in points], [p[1] for p in points], s=10, alpha=0.7)
            ax.set_xlabel(x)
            ax.set_ylabel(y)
            ax.set_title(title or f"Scatter: {x} vs {y}")
            fig.tight_layout()
            fig.savefig(out_path)
            plt.close(fig)
        except Exception as e:
            return f"Error: Failed to save plot: {type(e).__name__}: {e}"

        return f"Saved scatter plot to {path}."


class SaveBarPlotTool(BaseTool):
    REQUIRES_CONFIRMATION = True
    OPERATION_TYPE = "write"
    CONFIRMATION_CHECK_ARG = "path"
    CONFIRMATION_MESSAGE = "Save bar plot to '{path}'"

    @property
    def name(self) -> str:
        return "save_bar_plot"

    @property
    def description(self) -> str:
        return (
            "Save a bar plot of the top value counts for a column to an image file (png/svg). "
            "Ask the user for the output path before calling."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataset_id": {"type": "string", "description": "Dataset ID."},
                "column": {"type": "string", "description": "Column name."},
                "path": {"type": "string", "description": "Output image path."},
                "top_n": {"type": "integer", "description": "Number of categories to plot.", "default": 20},
                "title": {"type": "string", "description": "Optional plot title."},
                "overwrite": {"type": "boolean", "description": "Overwrite if file exists.", "default": False},
            },
            "required": ["dataset_id", "column", "path"],
        }

    def execute(
        self,
        dataset_id: str,
        column: str,
        path: str,
        top_n: int = 20,
        title: str | None = None,
        overwrite: bool = False,
    ) -> str:
        ds = _get_dataset(dataset_id)
        if not ds:
            return f"Error: Unknown dataset_id '{dataset_id}'."
        if column not in ds.columns:
            return f"Error: Unknown column '{column}'. Available: {', '.join(ds.columns)}"

        try:
            out_path = _resolve_path(path)
        except PathTraversalError as e:
            return f"Security error: {e}"

        if out_path.exists() and not overwrite:
            return (
                f"Error: Output file already exists: {path}. "
                "Ask the user whether to overwrite, then retry with overwrite=true."
            )

        counts: dict[str, int] = {}
        for r in ds.rows:
            v = r.get(column)
            if _is_missing(v):
                continue
            key = _format_value(v)
            counts[key] = counts.get(key, 0) + 1
        items = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)[: max(1, min(int(top_n), 50))]
        if not items:
            return f"Error: Column '{column}' has no non-missing values."

        labels = [k for k, _ in items]
        values = [v for _, v in items]

        try:
            import matplotlib  # type: ignore[import-not-found]

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt  # type: ignore[import-not-found]
        except Exception:
            return "Error: matplotlib is required for plot saving. Install with: uv sync --extra data"

        try:
            _ensure_parent_dir(out_path)
            fig, ax = plt.subplots(figsize=(max(6, 0.35 * len(labels)), 5))
            ax.bar(labels, values)
            ax.set_xlabel(column)
            ax.set_ylabel("count")
            ax.set_title(title or f"Top {len(labels)}: {column}")
            ax.tick_params(axis="x", rotation=45, labelsize=9)
            fig.tight_layout()
            fig.savefig(out_path)
            plt.close(fig)
        except Exception as e:
            return f"Error: Failed to save plot: {type(e).__name__}: {e}"

        return f"Saved bar plot to {path}."
